import openpyxl
import datetime
import csv
import json
import argparse
import os

def convert_excel_to_csv(excel_path, output_csv_path, dow_format='full', delimiter=';', replace_v=True, output_js_path='cuadrante_data.js'):
    """
    Convierte la hoja de proyección del cuadrante de turnos (2005-2040) a un archivo CSV y a cuadrante_data.js.
    """
    if not os.path.exists(excel_path):
        raise FileNotFoundError(f"No se encontró el archivo Excel: {excel_path}")

    wb = openpyxl.load_workbook(excel_path, data_only=True)
    
    # Buscar la pestaña 'Año 2009a2040'
    sheet_name = None
    for name in wb.sheetnames:
        if '2009' in name or '2040' in name:
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[1]
        
    ws = wb[sheet_name]
    print(f"Procesando pestaña: '{ws.title}' ({ws.max_row} filas, {ws.max_column} columnas)")

    dow_full_es = ['Lunes', 'Martes', 'Miércoles', 'Jueves', 'Viernes', 'Sábado', 'Domingo']
    dow_initial_es = ['L', 'M', 'X', 'J', 'V', 'S', 'D']
    dow_short_es = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']

    rows = []
    # Cabecera del CSV
    rows.append(['Fecha', 'Equipo A', 'Equipo B', 'Equipo C', 'Equipo D', 'Equipo E', 'Equipo F'])
    
    js_data = {}

    count_days = 0
    v_replaced_o = 0
    v_replaced_d = 0

    # Cada año ocupa exactamente 97 filas (Fila 1: Año, Filas 2-97: 12 meses x 8 filas)
    for year_row in range(1, ws.max_row + 1, 97):
        year_val = ws.cell(year_row, 1).value
        if year_val is None:
            break
        year = int(year_val)
        
        for m in range(1, 13):
            m_start = year_row + 1 + (m - 1) * 8
            
            for col in range(1, 32):
                day_num = ws.cell(m_start + 1, col).value
                if day_num is not None and isinstance(day_num, (int, float)):
                    day_num = int(day_num)
                    dt = datetime.date(year, m, day_num)
                    
                    w = dt.weekday()
                    if dow_format == 'initial':
                        dow_str = dow_initial_es[w]
                    elif dow_format == 'short':
                        dow_str = dow_short_es[w]
                    else: # 'full'
                        dow_str = dow_full_es[w]
                        
                    fecha_str = f"{dow_str} {dt.strftime('%d/%m/%Y')}"
                    
                    shifts = []
                    for s in range(6):
                        val = ws.cell(m_start + 2 + s, col).value
                        val_str = str(val).strip() if val is not None else ''
                        
                        if replace_v and val_str == 'V':
                            if w < 5:  # L-V -> O
                                val_str = 'O'
                                v_replaced_o += 1
                            else:  # S-D -> D
                                val_str = 'D'
                                v_replaced_d += 1
                                
                        shifts.append(val_str)
                    
                    rows.append([fecha_str] + shifts)
                    
                    # Key YYYY-MM-DD para JS
                    key_js = dt.strftime('%Y-%m-%d')
                    js_data[key_js] = shifts
                    
                    count_days += 1

    # Guardar CSV con codificación utf-8-sig (BOM)
    with open(output_csv_path, 'w', newline='', encoding='utf-8-sig') as f:
        writer = csv.writer(f, delimiter=delimiter)
        writer.writerows(rows)

    # Guardar JS bundle para carga síncrona sin problemas de CORS local (file://)
    if output_js_path:
        with open(output_js_path, 'w', encoding='utf-8') as f:
            f.write('window.CUADRANTE_DATA = ')
            json.dump(js_data, f, separators=(',', ':'))
            f.write(';')

    print(f"¡Exportación finalizada!")
    print(f"- Días procesados: {count_days}")
    if replace_v:
        print(f"- Reemplazos de 'V': {v_replaced_o} por 'O' (L-V), {v_replaced_d} por 'D' (S-D)")
    print(f"- CSV generado: {os.path.abspath(output_csv_path)}")
    if output_js_path:
        print(f"- JS Data generado: {os.path.abspath(output_js_path)}")

if __name__ == '__main__':
    parser = argparse.ArgumentParser(description="Convertidor de Cuadrante de Turnos a CSV y JS")
    parser.add_argument('--excel', default='Ciclo trabajo casi perpetuo (hasta 2040).xlsm', help="Ruta al archivo Excel")
    parser.add_argument('--output', default='cuadrante_perpetuo_iniciales.csv', help="Ruta de salida del CSV")
    parser.add_argument('--format', choices=['full', 'initial', 'short'], default='initial', help="Formato día semana")
    parser.add_argument('--delimiter', default=';', help="Delimitador CSV")
    parser.add_argument('--no-replace-v', action='store_true', help="Desactivar reemplazo de V")
    
    args = parser.parse_args()
    convert_excel_to_csv(args.excel, args.output, dow_format=args.format, delimiter=args.delimiter, replace_v=not args.no_replace_v)
